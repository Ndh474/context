import json
import openpyxl
import argparse
import sys
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def process_single_scenario(data, wb, add_to_file=None):
    """
    Xử lý một scenario duy nhất và tạo/thêm sheet vào workbook

    Args:
        data: Dictionary chứa dữ liệu của một scenario
        wb: Workbook để thêm sheet vào
        add_to_file: Đường dẫn file Excel (dùng để xác định tạo mới hay thêm vào)

    Returns:
        sheet_name, output_file, statistics
    """
    metadata = data.get("metadata", {})
    statistics = data.get("statistics", {})
    preconditions = data.get("preconditions", [])
    inputs = data.get("inputs", [])
    confirm_groups = data.get("confirm_groups", [])
    test_cases = data.get("test_cases", [])

    # Lấy sheet_name từ metadata (mặc định: "Test Cases")
    sheet_name = metadata.get("sheet_name", "Test Cases")

    # Kiểm tra nếu sheet đã tồn tại thì xóa đi
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Tạo sheet mới
    if add_to_file or len(wb.sheetnames) > 1 or wb.sheetnames[0] != "Sheet":
        # Thêm sheet vào workbook có sẵn
        ws = wb.create_sheet(sheet_name)
    else:
        # Workbook mới, dùng sheet mặc định
        ws = wb.active
        ws.title = sheet_name

    # --- 1. CẤU HÌNH STYLE ---
    navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    white_bold_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    black_bold_font = Font(color="000000", bold=True, name="Calibri", size=10)
    black_normal_font = Font(color="000000", name="Calibri", size=10)
    thin_border = Side(border_style="thin", color="000000")

    def set_border(cell):
        cell.border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

    def render_preconditions(ws, preconditions, current_row):
        """
        Render preconditions (không có O, chỉ hiển thị description)

        Args:
            ws: worksheet
            preconditions: list of precondition items
            current_row: row hiện tại để bắt đầu render

        Returns:
            current_row sau khi render, start_row, end_row
        """
        if not preconditions:
            return current_row, current_row, current_row - 1

        start_row = current_row

        # Header row - "Precondition" bold
        ws.merge_cells(
            start_row=current_row, start_column=2, end_row=current_row, end_column=4
        )
        cell_header = ws.cell(row=current_row, column=2)
        cell_header.value = "Precondition"
        cell_header.font = black_bold_font
        cell_header.alignment = Alignment(horizontal="left", vertical="center")
        set_border(cell_header)
        current_row += 1

        # Data rows - chỉ hiển thị description, KHÔNG có O
        for item in preconditions:
            ws.merge_cells(
                start_row=current_row, start_column=2, end_row=current_row, end_column=4
            )
            cell_desc = ws.cell(row=current_row, column=2)
            cell_desc.value = item.get("description", "")
            cell_desc.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="right"
            )
            cell_desc.font = black_normal_font
            set_border(cell_desc)
            current_row += 1

        end_row = current_row - 1
        return current_row, start_row, end_row

    def render_group(ws, group, current_row, start_col, test_cases):
        """
        Render một group (có header và items với O) vào worksheet

        Args:
            ws: worksheet
            group: dict với {"header": str, "items": list}
            current_row: row hiện tại để bắt đầu render
            start_col: cột bắt đầu của test cases
            test_cases: danh sách test cases

        Returns:
            current_row sau khi render xong group, start_row, end_row
        """
        group_start_row = current_row

        # Header row - Bold, căn trái
        ws.merge_cells(
            start_row=current_row, start_column=2, end_row=current_row, end_column=4
        )
        cell_header = ws.cell(row=current_row, column=2)
        cell_header.value = group.get("header", "")
        cell_header.font = black_bold_font
        cell_header.alignment = Alignment(horizontal="left", vertical="center")
        set_border(cell_header)
        current_row += 1

        # Data rows - Căn phải, CÓ O
        for item in group.get("items", []):
            ws.merge_cells(
                start_row=current_row, start_column=2, end_row=current_row, end_column=4
            )
            cell_desc = ws.cell(row=current_row, column=2)
            cell_desc.value = item.get("description", "")
            cell_desc.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="right"
            )
            cell_desc.font = black_normal_font
            set_border(cell_desc)

            # Tick "O" vào các TC tương ứng
            applicable_to = item.get("applicable_to", [])
            for tc_idx, tc in enumerate(test_cases):
                col = start_col + tc_idx
                if tc.get("id") in applicable_to:
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = "O"
                    cell.font = black_bold_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    set_border(cell)

            current_row += 1

        group_end_row = current_row - 1
        return current_row, group_start_row, group_end_row

    # --- 2. SETUP HEADER (Row 1-3) ---
    ws["A1"] = "Code Module"
    ws["C1"] = metadata.get("code_module", "")
    ws["E1"] = "Method"
    ws["K1"] = metadata.get("method", "")
    ws["A2"] = "Created By"
    ws["C2"] = metadata.get("created_by", "")
    ws["E2"] = "Executed By"
    ws["K2"] = metadata.get("executed_by", "")
    ws["A3"] = "Test requirement"
    ws["C3"] = metadata.get("test_requirement", "")

    for cell in ["A1", "E1", "A2", "E2", "A3"]:
        ws[cell].font = black_bold_font

    # Merge cells cho header
    ws.merge_cells("A1:B1")
    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:B3")
    ws.merge_cells("C1:D1")
    ws.merge_cells("C2:D2")
    ws.merge_cells("E1:J1")
    ws.merge_cells("E2:J2")
    ws.merge_cells("K1:S1")
    ws.merge_cells("K2:S2")
    ws.merge_cells("C3:S3")

    # --- 3. STATISTICS (Row 4-5) ---
    ws.merge_cells("A4:B4")
    ws["A4"] = "Passed"
    ws.merge_cells("C4:D4")
    ws["C4"] = "Failed"
    ws.merge_cells("E4:J4")
    ws["E4"] = "Untested"
    ws.merge_cells("K4:M4")
    ws["K4"] = "N / A / B"
    ws.merge_cells("N4:S4")
    ws["N4"] = "Total Test Cases"

    # Fill statistics từ JSON
    ws.merge_cells("A5:B5")
    ws["A5"] = statistics.get("passed", 0)
    ws.merge_cells("C5:D5")
    ws["C5"] = statistics.get("failed", 0)
    ws.merge_cells("E5:J5")
    ws["E5"] = statistics.get("untested", 0)

    type_counts = statistics.get("type_counts", {})
    ws["K5"] = type_counts.get("N", 0)
    ws["L5"] = type_counts.get("A", 0)
    ws["M5"] = type_counts.get("B", 0)

    ws.merge_cells("N5:S5")
    ws["N5"] = statistics.get("total_test_cases", len(test_cases))

    for c in ["K5", "L5", "M5"]:
        ws[c].alignment = Alignment(horizontal="center")
        ws[c].font = black_normal_font

    # Style Row 4
    for cell in ws[4]:
        cell.font = black_bold_font
        cell.alignment = Alignment(horizontal="center")

    # Style Row 5
    for cell in ws[5]:
        if cell.column not in [11, 12, 13]:  # K, L, M không merge
            cell.alignment = Alignment(horizontal="center")

    # Border cho header + statistics (row 1-5)
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=19):
        for cell in row:
            set_border(cell)

    # --- 4. MATRIX HEADER (Row 7) ---
    # Cột A-D: navy fill cho sidebar và description
    for col in range(1, 5):
        cell = ws.cell(row=7, column=col)
        cell.fill = navy_fill
        set_border(cell)

    # Merge cột B-D cho phần description header
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)

    # Cột E trở đi: TC headers (TC001, TC002, ...) - Text xoay dọc 90°
    start_col = 5  # Cột E
    for idx, tc in enumerate(test_cases):
        col = start_col + idx
        cell = ws.cell(row=7, column=col)
        cell.value = tc.get("id", f"TC{idx+1:03d}")
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=90
        )
        set_border(cell)

    # --- 5. MATRIX CONTENT - CONDITION SECTION ---
    current_row = 8
    condition_start_row = current_row

    # 5.1. Render Preconditions (KHÔNG có O)
    current_row, precond_start, precond_end = render_preconditions(
        ws, preconditions, current_row
    )

    # 5.2. Render Input groups (CÓ O)
    for group in inputs:
        current_row, group_start, group_end = render_group(
            ws, group, current_row, start_col, test_cases
        )

    condition_end_row = current_row - 1

    # --- 6. MATRIX CONTENT - CONFIRM GROUPS ---
    confirm_start_row = current_row

    for group in confirm_groups:
        current_row, group_start, group_end = render_group(
            ws, group, current_row, start_col, test_cases
        )

    confirm_end_row = current_row - 1

    # --- 7. MATRIX CONTENT - RESULT ROWS ---
    result_start_row = current_row

    # Row: Type(N : Normal, A : Abnormal, B : Boundary)
    ws.merge_cells(
        start_row=current_row, start_column=2, end_row=current_row, end_column=4
    )
    cell_b = ws.cell(row=current_row, column=2)
    cell_b.value = "Type(N : Normal, A : Abnormal, B : Boundary)"
    cell_b.font = black_bold_font
    cell_b.alignment = Alignment(wrap_text=True, vertical="center")

    for idx, tc in enumerate(test_cases):
        col = start_col + idx
        cell = ws.cell(row=current_row, column=col)
        cell.value = tc.get("type", "")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        set_border(cell)

    current_row += 1

    # Row: Passed/Failed
    ws.merge_cells(
        start_row=current_row, start_column=2, end_row=current_row, end_column=4
    )
    cell_b = ws.cell(row=current_row, column=2)
    cell_b.value = "Passed/Failed"
    cell_b.font = black_bold_font
    cell_b.alignment = Alignment(wrap_text=True, vertical="center")

    for idx, tc in enumerate(test_cases):
        col = start_col + idx
        cell = ws.cell(row=current_row, column=col)
        cell.value = tc.get("passed_failed", "")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        set_border(cell)

    current_row += 1

    # Row: Executed Date - Gộp thành 1 cột duy nhất
    ws.merge_cells(
        start_row=current_row, start_column=2, end_row=current_row, end_column=4
    )
    cell_b = ws.cell(row=current_row, column=2)
    cell_b.value = "Executed Date"
    cell_b.font = black_bold_font
    cell_b.alignment = Alignment(wrap_text=True, vertical="center")

    # Merge tất cả các cột TC thành 1 cột duy nhất cho date
    ws.merge_cells(
        start_row=current_row,
        start_column=start_col,
        end_row=current_row,
        end_column=start_col + len(test_cases) - 1,
    )
    cell_date = ws.cell(row=current_row, column=start_col)
    # Lấy executed_date từ metadata
    cell_date.value = metadata.get("executed_date", "")
    cell_date.alignment = Alignment(horizontal="center", vertical="center")
    set_border(cell_date)

    current_row += 1

    # Row: Defect ID
    ws.merge_cells(
        start_row=current_row, start_column=2, end_row=current_row, end_column=4
    )
    cell_b = ws.cell(row=current_row, column=2)
    cell_b.value = "Defect ID"
    cell_b.font = black_bold_font
    cell_b.alignment = Alignment(wrap_text=True, vertical="center")

    for idx, tc in enumerate(test_cases):
        col = start_col + idx
        cell = ws.cell(row=current_row, column=col)
        cell.value = tc.get("defect_id", "")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        set_border(cell)

    result_end_row = current_row

    # --- 8. SIDE BAR (CỘT A) ---
    # Condition section - Set value TRƯỚC KHI merge
    if condition_start_row <= condition_end_row:
        cell = ws.cell(row=condition_start_row, column=1)
        cell.value = "Condition"
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="left", vertical="top", indent=1)

        ws.merge_cells(
            start_row=condition_start_row,
            start_column=1,
            end_row=condition_end_row,
            end_column=1,
        )

    # Confirm section - Set value TRƯỚC KHI merge
    if confirm_start_row <= confirm_end_row:
        cell = ws.cell(row=confirm_start_row, column=1)
        cell.value = "Confirm"
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="left", vertical="top", indent=1)

        ws.merge_cells(
            start_row=confirm_start_row,
            start_column=1,
            end_row=confirm_end_row,
            end_column=1,
        )

    # Result section - Set value TRƯỚC KHI merge
    cell = ws.cell(row=result_start_row, column=1)
    cell.value = "Result"
    cell.fill = navy_fill
    cell.font = white_bold_font
    cell.alignment = Alignment(horizontal="left", vertical="top", indent=1)

    ws.merge_cells(
        start_row=result_start_row, start_column=1, end_row=result_end_row, end_column=1
    )

    # --- 9. BORDER CHO TOÀN BỘ MATRIX ---
    # Border cho cột A-D (sidebar + description)
    for row in ws.iter_rows(min_row=7, max_row=result_end_row, min_col=1, max_col=4):
        for cell in row:
            set_border(cell)

    # Border cho cột test cases (E trở đi)
    max_tc_col = start_col + len(test_cases) - 1
    for row in ws.iter_rows(
        min_row=7, max_row=result_end_row, min_col=start_col, max_col=max_tc_col
    ):
        for cell in row:
            set_border(cell)

    # --- 10. ĐIỀU CHỈNH WIDTH ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # Test case columns: mỗi cột rộng 8
    for idx in range(len(test_cases)):
        col = start_col + idx
        ws.column_dimensions[get_column_letter(col)].width = 8

    # Calculate statistics
    total_input_items = sum(len(g.get("items", [])) for g in inputs)
    total_confirm_items = sum(len(g.get("items", [])) for g in confirm_groups)

    return {
        "sheet_name": sheet_name,
        "test_cases_count": len(test_cases),
        "preconditions_count": len(preconditions),
        "input_groups_count": len(inputs),
        "input_items_count": total_input_items,
        "confirm_groups_count": len(confirm_groups),
        "confirm_items_count": total_confirm_items,
    }


def create_final_excel_from_json(
    json_file="test_data.json",
    add_to_file=None,
):
    """
    Tạo Excel test case từ file JSON (hỗ trợ cả single object và array)

    Args:
        json_file: Đường dẫn file JSON input
        add_to_file: Đường dẫn file Excel để thêm sheet vào (nếu None sẽ tạo mới)
    """
    # Load JSON data
    with open(json_file, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    # Kiểm tra xem JSON là single object hay array
    if isinstance(json_data, list):
        scenarios = json_data
    else:
        scenarios = [json_data]

    # Tạo hoặc mở workbook
    if add_to_file:
        wb = openpyxl.load_workbook(add_to_file)
    else:
        wb = openpyxl.Workbook()

    # Xử lý từng scenario
    results = []
    for scenario in scenarios:
        result = process_single_scenario(scenario, wb, add_to_file)
        results.append(result)

    # Save file
    if add_to_file:
        wb.save(add_to_file)
        output_file = add_to_file
        action = "added to"
    else:
        # Tạo tên file output từ JSON name
        output_file = json_file.replace(".json", "_output.xlsx")
        wb.save(output_file)
        action = "created"

    # Print results
    if len(results) == 1:
        # Single scenario
        result = results[0]
        print(
            f"[OK] Sheet '{result['sheet_name']}' {action} file '{output_file}' successfully!"
        )
        print(f"     Test Cases: {result['test_cases_count']}")
        print(f"     Preconditions: {result['preconditions_count']} items (no O)")
        print(
            f"     Input Groups: {result['input_groups_count']} ({result['input_items_count']} items with O)"
        )
        print(
            f"     Confirm Groups: {result['confirm_groups_count']} ({result['confirm_items_count']} items)"
        )
    else:
        # Multiple scenarios
        print(f"[OK] {len(results)} sheets {action} file '{output_file}' successfully!")
        for result in results:
            print(
                f"  - {result['sheet_name']}: {result['test_cases_count']} test cases"
            )

    return output_file


def main():
    """Main CLI entry point"""
    parser = argparse.ArgumentParser(
        description="Generate Excel test case from JSON data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Create new Excel file from JSON
  python main.py --json test_data.json
  
  # Add sheet to existing Excel file
  python main.py --json test_data.json --add-sheet existing.xlsx
        """,
    )

    parser.add_argument("--json", required=True, help="Path to JSON input file")

    parser.add_argument(
        "--add-sheet",
        help="Add sheet to existing Excel file (if not specified, creates new file)",
    )

    args = parser.parse_args()

    try:
        create_final_excel_from_json(json_file=args.json, add_to_file=args.add_sheet)
    except FileNotFoundError as e:
        print(f"[ERROR] File not found: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
