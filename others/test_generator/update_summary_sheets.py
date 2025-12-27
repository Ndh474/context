"""
Update MethodList and Statistics sheets in Excel from JSON test data.

Usage:
    python update_summary_sheets.py --json-dir test_json --excel test_output/Report5.1_UnitTest.xlsx
"""

import argparse
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def load_all_json_data(json_dir: str) -> list:
    """Load all JSON files and flatten into list of scenarios."""
    all_scenarios = []
    
    for filename in sorted(os.listdir(json_dir)):
        if filename.endswith('.json'):
            filepath = os.path.join(json_dir, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    all_scenarios.extend(data)
                else:
                    all_scenarios.append(data)
    
    return all_scenarios


def update_method_list(wb, scenarios: list):
    """Update MethodList sheet with data from all scenarios."""
    ws = wb['MethodList']
    
    # Styles
    thin_border = Side(border_style="thin", color="000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    normal_font = Font(name="Calibri", size=10)
    no_fill = PatternFill()  # White/no fill
    
    # Clear header row 10 columns E-F (Description, Pre-Condition)
    for col in [5, 6]:
        cell = ws.cell(row=10, column=col)
        cell.value = None
        cell.fill = no_fill
        cell.border = Border()
    
    # Data starts from row 11
    start_row = 11
    
    # Clear existing data and format (rows 11 onwards), columns A-F
    for row in range(start_row, 300):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = Border()
    
    # Fill data - only 4 columns: No, Module Name, Method Name, Sheet Name
    row_num = start_row
    for idx, scenario in enumerate(scenarios, 1):
        metadata = scenario.get('metadata', {})
        
        # No (column A)
        ws.cell(row=row_num, column=1).value = idx
        ws.cell(row=row_num, column=1).border = border
        ws.cell(row=row_num, column=1).font = normal_font
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        
        # Module Name (column B)
        ws.cell(row=row_num, column=2).value = metadata.get('code_module', '')
        ws.cell(row=row_num, column=2).border = border
        ws.cell(row=row_num, column=2).font = normal_font
        
        # Method Name (column C)
        ws.cell(row=row_num, column=3).value = metadata.get('method', '')
        ws.cell(row=row_num, column=3).border = border
        ws.cell(row=row_num, column=3).font = normal_font
        
        # Sheet Name as hyperlink (column D)
        sheet_name = metadata.get('sheet_name', '')
        ws.cell(row=row_num, column=4).value = sheet_name
        ws.cell(row=row_num, column=4).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row_num, column=4).font = Font(name="Calibri", size=10, color="0000FF", underline="single")
        ws.cell(row=row_num, column=4).border = border
        
        row_num += 1
    
    # Update total count in header (row 6, column C)
    ws['C6'] = len(scenarios)
    
    print(f"[OK] MethodList: {len(scenarios)} methods")


def update_statistics(wb, scenarios: list):
    """Update Statistics sheet with aggregated data."""
    ws = wb['Statistics']
    
    # Styles
    thin_border = Side(border_style="thin", color="000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    normal_font = Font(name="Calibri", size=10)
    navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    no_fill = PatternFill()
    
    # Header is at row 11, data starts from row 12
    start_row = 12
    
    # Clear existing data (from row 12 to 200) - set value to empty string first to clear formulas
    for row in range(start_row, 200):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.value = ""  # Clear any formulas
            cell.value = None
            cell.fill = no_fill
            cell.border = Border()
            cell.font = normal_font
            cell.number_format = 'General'
    
    # Totals
    total_passed = 0
    total_failed = 0
    total_untested = 0
    total_n = 0
    total_a = 0
    total_b = 0
    total_tc = 0
    
    # Fill data
    row_num = start_row
    for idx, scenario in enumerate(scenarios, 1):
        metadata = scenario.get('metadata', {})
        stats = scenario.get('statistics', {})
        type_counts = stats.get('type_counts', {})
        
        passed = stats.get('passed', 0)
        failed = stats.get('failed', 0)
        untested = stats.get('untested', 0)
        n_count = type_counts.get('N', 0)
        a_count = type_counts.get('A', 0)
        b_count = type_counts.get('B', 0)
        tc_count = stats.get('total_test_cases', 0)
        
        total_passed += passed
        total_failed += failed
        total_untested += untested
        total_n += n_count
        total_a += a_count
        total_b += b_count
        total_tc += tc_count
        
        # No (column A)
        ws.cell(row=row_num, column=1).value = idx
        ws.cell(row=row_num, column=1).border = border
        ws.cell(row=row_num, column=1).font = normal_font
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        
        # Function code - sheet name as hyperlink (column B)
        sheet_name = metadata.get('sheet_name', '')
        ws.cell(row=row_num, column=2).value = sheet_name
        ws.cell(row=row_num, column=2).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row_num, column=2).font = Font(name="Calibri", size=10, color="0000FF", underline="single")
        ws.cell(row=row_num, column=2).border = border
        
        # Passed (column C)
        ws.cell(row=row_num, column=3).value = passed
        ws.cell(row=row_num, column=3).border = border
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        
        # Failed (column D)
        ws.cell(row=row_num, column=4).value = failed
        ws.cell(row=row_num, column=4).border = border
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
        
        # Untested (column E)
        ws.cell(row=row_num, column=5).value = untested
        ws.cell(row=row_num, column=5).border = border
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
        
        # N (column F)
        ws.cell(row=row_num, column=6).value = n_count
        ws.cell(row=row_num, column=6).border = border
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        
        # A (column G)
        ws.cell(row=row_num, column=7).value = a_count
        ws.cell(row=row_num, column=7).border = border
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='center')
        
        # B (column H)
        ws.cell(row=row_num, column=8).value = b_count
        ws.cell(row=row_num, column=8).border = border
        ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='center')
        
        # Total Test Cases (column I)
        ws.cell(row=row_num, column=9).value = tc_count
        ws.cell(row=row_num, column=9).border = border
        ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')
        
        row_num += 1
    
    # Sub total row (right after last data row)
    sub_total_row = row_num
    
    # Column A - empty
    ws.cell(row=sub_total_row, column=1).value = ""
    ws.cell(row=sub_total_row, column=1).border = border
    
    # Column B - "Sub total" with navy fill
    ws.cell(row=sub_total_row, column=2).value = "Sub total"
    ws.cell(row=sub_total_row, column=2).fill = navy_fill
    ws.cell(row=sub_total_row, column=2).font = white_font
    ws.cell(row=sub_total_row, column=2).border = border
    
    # Columns C-I - totals with navy fill
    totals = [total_passed, total_failed, total_untested, total_n, total_a, total_b, total_tc]
    for col_idx, val in enumerate(totals, 3):
        cell = ws.cell(row=sub_total_row, column=col_idx)
        cell.value = val
        cell.fill = navy_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Summary stats section (2 rows after sub total)
    summary_start = sub_total_row + 2
    
    # Test coverage = (passed + failed) / total * 100
    coverage = ((total_passed + total_failed) / total_tc * 100) if total_tc > 0 else 0
    # Test successful coverage = passed / total * 100  
    success_coverage = (total_passed / total_tc * 100) if total_tc > 0 else 0
    # Normal case %
    normal_pct = (total_n / total_tc * 100) if total_tc > 0 else 0
    # Abnormal case %
    abnormal_pct = (total_a / total_tc * 100) if total_tc > 0 else 0
    # Boundary case %
    boundary_pct = (total_b / total_tc * 100) if total_tc > 0 else 0
    
    # Write summary labels and values
    summary_data = [
        ("Test coverage", f"{coverage:.1f} %"),
        ("Test successful coverage", f"{success_coverage:.1f} %"),
        ("Normal case", f"{normal_pct:.1f} %"),
        ("Abnormal case", f"{abnormal_pct:.1f} %"),
        ("Boundary case", f"{boundary_pct:.1f} %"),
    ]
    
    for i, (label, value) in enumerate(summary_data):
        row = summary_start + i
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = normal_font
        ws.cell(row=row, column=4).value = value
        ws.cell(row=row, column=4).font = normal_font
    
    print(f"[OK] Statistics: {len(scenarios)} sheets, {total_tc} TCs")
    print(f"     Passed: {total_passed}, Failed: {total_failed}, Untested: {total_untested}")
    print(f"     N: {total_n}, A: {total_a}, B: {total_b}")
    print(f"     Coverage: {coverage:.1f}%")


def main():
    parser = argparse.ArgumentParser(
        description='Update MethodList and Statistics sheets from JSON data'
    )
    parser.add_argument('--json-dir', default='test_json', help='Directory containing JSON files')
    parser.add_argument('--excel', default='test_output/Report5.1_UnitTest.xlsx', help='Excel file to update')
    
    args = parser.parse_args()
    
    if not os.path.isdir(args.json_dir):
        print(f"[ERROR] JSON directory not found: {args.json_dir}")
        return 1
    
    if not os.path.exists(args.excel):
        print(f"[ERROR] Excel file not found: {args.excel}")
        return 1
    
    # Load all JSON data
    print(f"Loading JSON files from: {args.json_dir}")
    scenarios = load_all_json_data(args.json_dir)
    print(f"Found {len(scenarios)} test scenarios")
    
    # Open Excel
    wb = openpyxl.load_workbook(args.excel)
    
    # Update sheets
    update_method_list(wb, scenarios)
    update_statistics(wb, scenarios)
    
    # Save
    wb.save(args.excel)
    print(f"\n[OK] Saved: {args.excel}")
    
    return 0


if __name__ == '__main__':
    exit(main())
